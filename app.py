import os
import tempfile
from flask import Flask, request, render_template, send_file, redirect, url_for, flash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
from zipfile import ZipFile
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Replace with your secret key
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Grading Scheme
def calculate_grade(percentage):
    try:
        percentage = float(percentage)
    except (ValueError, TypeError):
        return 'Invalid', 0.00
    if percentage >= 90:
        return 'A+', 4.00
    elif percentage >= 85:
        return 'A', 3.67
    elif percentage >= 80:
        return 'B+', 3.33
    elif percentage >= 75:
        return 'B', 3.00
    elif percentage >= 70:
        return 'C+', 2.67
    elif percentage >= 60:
        return 'C', 2.33
    else:
        return 'F', 0.00

# Calculate SGPA for a student
def calculate_sgpa(grades):
    total_points = 0
    total_credits = 0
    for grade in grades:
        try:
            credits = float(grade['credits'])
        except (ValueError, TypeError):
            credits = 0.0
        grade_points = grade['grade_points']
        total_points += grade_points * credits
        total_credits += credits
    return round(total_points / total_credits, 2) if total_credits > 0 else 0.00

# Allowed file check
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Process the uploaded Excel file
def process_excel_file(excel_file_path, temp_dir):
    """Process the uploaded Excel file and generate transcripts."""
    wb = load_workbook(excel_file_path, data_only=True)
    sheet = wb.active

    students = extract_students(sheet)

    # Set up the template environment
    env = Environment(loader=FileSystemLoader('templates'))
    template = env.get_template('transcript_template.html')

    pdf_files = []
    for student in students:
        # Calculate grades and GPA
        all_grades = []
        for course in student['courses']:
            # Use the course's total percentage to calculate grade
            letter_grade, grade_points = calculate_grade(course['total_percentage'])
            course['letter_grade'] = letter_grade
            course['grade_points'] = grade_points
            course['credits'] = 3  # Adjust credits as needed

            grade_entry = {
                'credits': course['credits'],
                'grade_points': course['grade_points']
            }
            all_grades.append(grade_entry)

        student['sgpa'] = calculate_sgpa(all_grades)
        student['cgpa'] = student['sgpa']  # Assuming SGPA and CGPA are the same here

        # Ensure student['name'] is a valid string for filename
        student_name = str(student.get('name', 'Unknown_Student'))  # Convert to string if not already
        student_name = student_name if student_name else 'Unknown_Student'

        # Render HTML
        html_out = template.render(student=student)

        # Generate PDF
        pdf_filename = f"{secure_filename(student_name)}_Transcript.pdf"
        pdf_file_path = os.path.join(temp_dir, pdf_filename)
        HTML(string=html_out).write_pdf(pdf_file_path)
        pdf_files.append((pdf_filename, pdf_file_path))

    return pdf_files

# Extract students from Excel sheet
def extract_students(sheet):
    """Extract student data from Excel sheet."""
    students = []
    courses_data = {}
    current_course_name = None

    # Read courses and components
    for col_idx in range(3, sheet.max_column + 1):  # Start from column C (index 3)
        course_name = sheet.cell(row=1, column=col_idx).value
        if course_name is not None:
            current_course_name = course_name
        else:
            course_name = current_course_name

        if course_name is None:
            continue  # Skip if course_name is still None

        component_name = sheet.cell(row=2, column=col_idx).value
        max_mark = sheet.cell(row=3, column=col_idx).value

        if course_name not in courses_data:
            courses_data[course_name] = {
                'components': [],
                'columns': [],
                'total_percentage_col': None
            }

        courses_data[course_name]['components'].append({
            'column_index': col_idx,
            'component_name': component_name,
            'max_mark': max_mark
        })
        courses_data[course_name]['columns'].append(col_idx)

    # After reading all columns, set total_percentage_col for each course
    for course_name, data in courses_data.items():
        if data['columns']:
            total_percentage_col = data['columns'][-1]  # Assuming last column is total percentage
            data['total_percentage_col'] = total_percentage_col

    # Read student data
    for row_idx in range(4, sheet.max_row + 1):
        serial_no = sheet.cell(row=row_idx, column=1).value  # Column A
        student_name = sheet.cell(row=row_idx, column=2).value  # Column B

        if not student_name:
            continue  # Skip rows without student name

        student = {
            'serial_no': serial_no,
            'name': student_name,
            'courses': []
        }

        for course_name, course_info in courses_data.items():
            course = {
                'name': course_name,
                'grades': []
            }
            total_marks_obtained = 0
            total_max_marks = 0
            for component in course_info['components']:
                col_idx = component['column_index']
                component_name = component['component_name']
                max_mark = component['max_mark']
                mark_obtained = sheet.cell(row=row_idx, column=col_idx).value

                if mark_obtained is None:
                    mark_obtained = 0

                # Keep track of total marks
                try:
                    total_marks_obtained += float(mark_obtained)
                    total_max_marks += float(max_mark)
                except (ValueError, TypeError):
                    pass  # Skip if marks are not numbers

                grade_entry = {
                    'subcomponent': component_name,
                    'percentage': mark_obtained,
                    'max_mark': max_mark,
                    'credits': 1  # Adjust as necessary
                }
                course['grades'].append(grade_entry)

            # Read total percentage from total_percentage_col
            total_percentage_col = course_info['total_percentage_col']
            if total_percentage_col:
                total_percentage = sheet.cell(row=row_idx, column=total_percentage_col).value
                if total_percentage is None:
                    total_percentage = 0
            else:
                # Calculate total percentage
                if total_max_marks > 0:
                    total_percentage = (total_marks_obtained / total_max_marks) * 100
                else:
                    total_percentage = 0

            course['total_percentage'] = total_percentage
            student['courses'].append(course)

        students.append(student)

    return students

# Create a zip file of PDFs
def create_zip_file(pdf_files, temp_dir):
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zipf:
        for filename, filepath in pdf_files:
            zipf.write(filepath, arcname=filename)
    zip_buffer.seek(0)
    return zip_buffer

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Process the file
        with tempfile.TemporaryDirectory() as temp_dir:
            pdf_files = process_excel_file(file_path, temp_dir)
            if not pdf_files:
                flash('No transcripts generated.')
                return redirect(url_for('index'))

            # Create zip
            zip_buffer = create_zip_file(pdf_files, temp_dir)
            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name='transcripts.zip'
            )
    else:
        flash('Allowed file types are .xlsx, .xls')
        return redirect(request.url)

if __name__ == '__main__':
    app.run(debug=True)
